"""
Garment Spec Sheet OCR Extractor — Streamlit App
Wraps original extraction logic. No changes to OCR code.
"""

import streamlit as st
import os, re, json, base64, requests, io, time
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance

# ── Config ───────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Garment Spec OCR",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

API_KEY = os.getenv("MISTRAL_API_KEY", "o5RwmxtgR18zW6ppUffmrSrEPRdEay39")
MODEL = "pixtral-large-2411"
MAX_TOKENS = 16000
TIMEOUT_SEC = 300
MAX_RETRIES = 3

OUTPUT_DIR = Path("./outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Prompts ──────────────────────────────────────────────────────────────────

PROMPT_FULL = """You are extracting data from a garment spec sheet photo. Output ONE row per measurement.

=== TABLE STRUCTURE ===

Col 1: Description
Col 2: Tolerance — two stacked cells show "Pos(+) X" and "Neg(-) X". Output as "+-X" e.g. "+-1", "+-0.5"
Then for each size group (read ALL from header): 3 sub-columns → BASE | ADJ | S1

Each measurement has 2 physical rows in the image (Pos on top, Neg below).
Output ONLY ONE row per measurement.

=== CELL MERGE RULE (strict priority) ===

For each ADJ and S1 cell, look at BOTH physical rows (Pos top + Neg bottom):
  1. Pos row has a number  AND  Neg row has "/"  → use the Pos number
  2. Neg row has a number  AND  Pos row has "/"  → use the Neg number
  3. BOTH rows have numbers                       → use the Pos row value
  4. BOTH rows are "/" or blank or checkmark      → output "/"

BASE column: printed number (same in both rows). Output once.
A checkmark (✓) always counts as "/".

=== BLANK CELL RULE (critical) ===

Any ADJ or S1 cell that is empty, contains only a checkmark (✓), or contains a forward slash (/)
MUST be output as exactly "/" — never output "-", "", or null.
There is NO dash character "-" used alone as a value in this table.
If you are tempted to write "-" for a cell, write "/" instead.

=== BASE VALUE VERIFICATION (critical — prevents column drift) ===

BASE values are pre-printed numbers that ALWAYS increase left to right across size groups.
Example for a measurement: 33.5 → 35 → 36.5 → 38 → 40 → 40.5 → 41.5

Before writing each BASE value, ask yourself:
  - Is it LARGER than the previous size group's BASE for this same measurement?
  - Does it match the printed number visible in THIS column (not the column to the left)?

COLUMN DRIFT WARNING: The most common error is copying the previous size group's BASE
instead of reading the new one, which shifts all values left by one column.
If you find yourself writing the same BASE twice in a row for the same measurement,
STOP — re-read that column from the top header down to the current row.

=== HANDWRITTEN NUMBER DECODING ===

Numbers use shorthand — NO decimal point is written:
  -1  = -1      +1  = +1
  -05 = -0.5    +05 = +0.5
  -03 = -0.3    +03 = +0.3
  -02 = -0.2    +02 = +0.2
  -04 = -0.4    +04 = +0.4
  -06 = -0.6    +06 = +0.6
  -07 = -0.7    +07 = +0.7
  -15 = -1.5    +15 = +1.5
  -2  = -2      +2  = +2

SIGN RULE: The sign written to the LEFT of digits ALWAYS belongs to that number.
  "-05" → -0.5 (NEVER +0.5).  "+05" → +0.5.

=== CIRCLED NUMBERS (critical — sign is OUTSIDE the circle) ===

A digit inside a drawn circle/oval IS a real value, NOT decoration.
The sign (+/-) is written OUTSIDE the circle. Look for it BEFORE assigning a sign.

  Circle "1"   + sign "+" written outside  → +1
  Circle "1"   + no sign outside           → +1   (no sign = positive)
  Circle "1"   + sign "-" written outside  → -1
  Circle "1.5" + sign "+" written outside  → +1.5
  Circle "1.5" + no sign outside           → +1.5  (no sign = positive)
  Circle "1.5" + sign "-" written outside  → -1.5
  Circle "05"  + no sign outside           → +0.5

NEVER output "/" for a circled number.
NEVER assign negative without seeing a "-" explicitly written outside the circle.

=== UNCERTAINTY RULE ===

If you are not sure whether an ADJ or S1 cell contains a value or is blank,
output "/" — do NOT guess a value. A false "/" is less harmful than a wrong number.

=== RIGHT COLUMNS — EXTRA ATTENTION ===

The last 2–3 size groups (5/6 YRS, 6/7 YRS, 7/8 YRS) are near the right edge.
Apply the BASE VERIFICATION rule especially carefully here.
Inspect every cell — faint or small handwriting still counts.

=== OUTPUT FORMAT ===

Return strict JSON only. No markdown. No explanation.

{
  "title": "<title from image>",
  "size_groups": ["1.5/2 YRS", "2/3 YRS", "3/4 YRS", "4/5 YRS", "5/6 YRS", "6/7 YRS", "7/8 YRS"],
  "rows": [
    ["A6 Length - SNP to Hem at Front", "+-1", "33.5", "-1", "-0.5", "35", "-0.5", "/", "36.5", "-0.5", "+1.5", "38", "-1", "/", "40", "-1", "+0.5", "40.5", "-1", "/", "41.5", "-0.5", "-1"],
    ... one row per measurement ...
  ]
}

Each row = 2 + (num_size_groups × 3) values.  For 7 size groups = 23 values per row.
Extract ALL measurements. Approximately 16–17 rows total."""


PROMPT_RIGHTCOLS = """You are re-checking the RIGHT SIDE of a garment spec sheet.
Focus ONLY on the last 3 size groups: 5/6 YRS, 6/7 YRS, 7/8 YRS.

=== RULES ===

CELL MERGE RULE (strict priority):
  1. Pos row has a number  AND  Neg row has "/"  → use the Pos number
  2. Neg row has a number  AND  Pos row has "/"  → use the Neg number
  3. BOTH rows have numbers                       → use Pos value
  4. BOTH "/" or blank or checkmark               → output "/"

BLANK CELL RULE: empty / checkmark / slash cells → ALWAYS output "/" — never "-".

HANDWRITTEN shorthand (no decimal written):
  -05=-0.5  +05=+0.5  -03=-0.3  +03=+0.3  -02=-0.2  +02=+0.2
  -07=-0.7  +07=+0.7  -15=-1.5  +15=+1.5  -1=-1  +1=+1

CIRCLED NUMBERS (sign is OUTSIDE the circle):
  Circle + "+" outside or no sign → positive (e.g. +1, +1.5, +0.5)
  Circle + "-" outside            → negative (e.g. -1, -1.5)
  NEVER output "/" for a circled number.
  NEVER assign negative without seeing an explicit "-" outside the circle.

BASE VERIFICATION: BASE values must increase left to right.
  5/6 BASE > 4/5 BASE, 6/7 BASE > 5/6 BASE, 7/8 BASE > 6/7 BASE.
  If a BASE looks the same as the previous group, re-read that column.

UNCERTAINTY RULE: when unsure about ADJ or S1 → output "/" not a guess.

=== OUTPUT FORMAT ===

Return strict JSON only. No markdown.
List every measurement row top to bottom. Output ONLY the 9 values for 5/6 + 6/7 + 7/8 YRS.

{
  "rows": [
    {"desc": "A6 Length - SNP to Hem at Front", "vals": ["40", "-1", "+0.5", "40.5", "-1", "/", "41.5", "-0.5", "-1"]},
    {"desc": "B1 Chest - 1cm from Underarm",    "vals": ["35", "-0.5", "/", "35.5", "-0.5", "/", "37.3", "/", "/"]},
    ... one entry per measurement row ...
  ]
}

Extract ALL ~16–17 measurement rows."""

# ── Original Logic (unchanged) ────────────────────────────────────────────────

def preprocess_image(image_bytes: bytes) -> tuple:
    """
    Open image, boost contrast + sharpness, encode as high-quality PNG.
    Returns (base64_str, mime_type, width, height).
    """
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    w, h = img.size

    longest = max(w, h)
    if longest < 1800:
        scale = 1800 / longest
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        w, h = img.size

    img = ImageEnhance.Contrast(img).enhance(1.4)
    img = ImageEnhance.Sharpness(img).enhance(1.5)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=False)
    b64 = base64.b64encode(buf.getvalue()).decode()

    return b64, "image/png", w, h


def call_api(api_key: str, b64: str, mime: str, prompt: str) -> dict:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": [
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
            {"type": "text", "text": prompt}
        ]}],
        "max_tokens": MAX_TOKENS,
        "temperature": 0.1,
    }
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.post("https://api.mistral.ai/v1/chat/completions",
                              headers=headers, json=payload, timeout=TIMEOUT_SEC)
            if r.status_code == 401:
                return {}
            if r.status_code == 429:
                time.sleep(30)
                continue
            if r.status_code != 200:
                return {}

            rj = r.json()
            choice = rj["choices"][0]
            finish = choice.get("finish_reason", "?")

            raw = choice["message"]["content"].strip()
            raw = re.sub(r'^```(?:json)?\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw).strip()
            if finish == "length":
                raw = _repair(raw)
            return _parse(raw)

        except requests.exceptions.ReadTimeout:
            if attempt < MAX_RETRIES:
                time.sleep(15 * attempt)
            else:
                return {}
        except requests.exceptions.ConnectionError:
            if attempt < MAX_RETRIES:
                time.sleep(10)
            else:
                return {}
    return {}


def _parse(raw: str) -> dict:
    for s in (raw, _repair(raw)):
        try:
            d = json.loads(s)
            if isinstance(d, dict) and ("rows" in d):
                return d
        except:
            pass
    return {}


def _repair(raw: str) -> str:
    last = raw.rfind('],')
    if last != -1:
        raw = raw[:last + 1]
    raw += ']' * max(raw.count('[') - raw.count(']'), 0)
    raw += '}' * max(raw.count('{') - raw.count('}'), 0)
    return raw


def normalize_slashes(result: dict) -> dict:
    """
    Fix #7: Replace any lone "-" or "" in ADJ/S1 columns with "/".
    """
    fixed = 0
    for row in result.get("rows", []):
        if not isinstance(row, list):
            continue
        for idx in range(3, len(row)):
            if (idx - 2) % 3 == 0:
                continue
            v = str(row[idx]).strip()
            if v in ("-", "", "—", "–"):
                row[idx] = "/"
                fixed += 1
    return result


def merge_passes(result1: dict, result2: dict) -> dict:
    """
    Overwrite last 9 values (5/6, 6/7, 7/8 YRS columns) of pass-1 rows
    with pass-2 values wherever pass-2 gives a non-'/' answer.
    """
    if not result2 or "rows" not in result2:
        return result1

    p2_map = {}
    for entry in result2["rows"]:
        if isinstance(entry, dict):
            desc = entry.get("desc", "").strip().lower()
            vals = entry.get("vals", [])
            if desc and len(vals) == 9:
                p2_map[desc] = vals

    merged = 0
    rows = result1.get("rows", [])
    for row in rows:
        if not isinstance(row, list) or len(row) < 23:
            continue
        desc_key = str(row[0]).strip().lower()
        p2 = p2_map.get(desc_key)
        if not p2:
            for k, v in p2_map.items():
                if k[:15] in desc_key or desc_key[:15] in k:
                    p2 = v
                    break
        if not p2:
            continue
        for i, v2 in enumerate(p2):
            idx = 14 + i
            if idx < len(row):
                v1 = str(row[idx])
                if v2 not in ("/", "", None) and v1 in ("/", "", None, "/"):
                    row[idx] = v2
                    merged += 1
                elif v2 not in ("/", "", None) and v1 != v2:
                    row[idx] = v2
                    merged += 1

    return result1


def border():
    s = Side(style='thin')
    return Border(top=s, bottom=s, left=s, right=s)

def sc(cell, bold=False, bg=None, ha='center', wrap=False, sz=9):
    cell.font = Font(name='Arial', bold=bold, size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border = border()
    if bg:
        cell.fill = PatternFill('solid', start_color=bg, end_color=bg)

HDR = "D9D9D9"
BASE_BG = "EBEBEB"

def write_excel(data: dict) -> bytes:
    rows = data.get("rows", [])
    title = data.get("title", "Specification")
    size_groups = data.get("size_groups", [])
    n_groups = len(size_groups)
    n_cols = 2 + n_groups * 3

    wb = Workbook()
    ws = wb.active
    ws.title = "Specifications"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name='Arial', bold=True, size=12)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    sc(ws.cell(row=2, column=1, value="Description"), bold=True, bg=HDR, ha='left', wrap=True, sz=8)
    sc(ws.cell(row=2, column=2, value="Tolerance"), bold=True, bg=HDR, wrap=True, sz=8)

    for gi, grp in enumerate(size_groups):
        bc = 3 + gi * 3
        ws.merge_cells(start_row=2, start_column=bc, end_row=2, end_column=bc + 2)
        for cc in range(bc, bc + 3):
            sc(ws.cell(row=2, column=cc), bold=True, bg=HDR, sz=8)
        ws.cell(row=2, column=bc).value = grp
    ws.row_dimensions[2].height = 22

    sc(ws.cell(row=3, column=1), bold=True, bg=HDR, sz=7)
    sc(ws.cell(row=3, column=2), bold=True, bg=HDR, sz=7)
    for gi in range(n_groups):
        bc = 3 + gi * 3
        for offset, lbl in enumerate(["BASE", "ADJ", "S1"]):
            sc(ws.cell(row=3, column=bc + offset, value=lbl), bold=True, bg=HDR, sz=7)
    ws.row_dimensions[3].height = 14

    for ri, row in enumerate(rows):
        er = ri + 4
        bg = "FFFFFF" if ri % 2 == 0 else "F2F2F2"
        rd = list(row) if isinstance(row, list) else []

        for ci in range(1, n_cols + 1):
            val = rd[ci - 1] if ci - 1 < len(rd) else ""
            cell = ws.cell(row=er, column=ci)
            cell.value = str(val) if val not in (None, "") else ""

            if ci == 1:
                sc(cell, bg=bg, ha='left', wrap=True, sz=8)
            elif ci == 2:
                sc(cell, bg=bg, ha='center', sz=9)
            else:
                col_within = (ci - 3) % 3
                sc(cell, bold=(col_within == 0),
                   bg=BASE_BG if col_within == 0 else bg,
                   sz=8 if col_within == 0 else 9)

        ws.row_dimensions[er].height = 16

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 7
    for gi in range(n_groups):
        bc = 3 + gi * 3
        ws.column_dimensions[get_column_letter(bc)].width = 7
        ws.column_dimensions[get_column_letter(bc + 1)].width = 7
        ws.column_dimensions[get_column_letter(bc + 2)].width = 7

    ws.freeze_panes = ws.cell(row=4, column=3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Streamlit UI ─────────────────────────────────────────────────────────────

st.markdown("""
<style>
    .main-title {
        font-size: 2.5rem;
        font-weight: 900;
        color: #1e293b;
        margin-bottom: 0.5rem;
    }
    .subtitle {
        font-size: 0.95rem;
        color: #64748b;
        margin-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="main-title">📊 Garment Spec OCR</p>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">Upload photo → Extract measurements → Download Excel</p>', unsafe_allow_html=True)

if not API_KEY.strip():
    st.error("❌ API key not configured. Set `MISTRAL_API_KEY` environment variable.")
    st.stop()

uploaded_file = st.file_uploader(
    "Upload garment spec photo",
    type=["jpg", "jpeg", "png", "bmp", "webp", "tiff"],
    help="Clear photo of measurement table"
)

if uploaded_file:
    col1, col2 = st.columns([3, 1])
    
    with col1:
        st.success(f"✓ {uploaded_file.name} ({uploaded_file.size // 1024} KB)")
    
    with col2:
        extract_btn = st.button("🚀 Extract", use_container_width=True, type="primary")
    
    if extract_btn:
        try:
            with st.spinner("📸 Preprocessing image..."):
                image_bytes = uploaded_file.getvalue()
                b64, mime, w, h = preprocess_image(image_bytes)
            
            with st.spinner("🔍 Pass 1: Extracting full table..."):
                result1 = call_api(API_KEY.strip(), b64, mime, PROMPT_FULL)
                
                if not result1 or not result1.get("rows"):
                    st.error("Pass-1 failed. Try clearer photo.")
                    st.stop()
                
                result1 = normalize_slashes(result1)
            
            with st.spinner("🔍 Pass 2: Re-checking right columns..."):
                result2 = call_api(API_KEY.strip(), b64, mime, PROMPT_RIGHTCOLS)
                if result2 and result2.get("rows"):
                    result1 = merge_passes(result1, result2)
                    result1 = normalize_slashes(result1)
            
            with st.spinner("📝 Generating Excel..."):
                excel_bytes = write_excel(result1)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"spec_{timestamp}.xlsx"
            
            st.success("✅ Extraction complete!")
            
            st.divider()
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Title", result1.get("title", "Spec")[:15] + "...")
            with col2:
                st.metric("Size Groups", len(result1.get("size_groups", [])))
            with col3:
                st.metric("Measurements", len(result1.get("rows", [])))
            
            st.divider()
            
            st.subheader("Size Groups")
            st.write(" • ".join(result1.get("size_groups", [])))
            
            st.subheader("Sample Measurements")
            sample_data = []
            for row in result1.get("rows", [])[:5]:
                if isinstance(row, list) and len(row) > 1:
                    sample_data.append({"Description": row[0], "Tolerance": row[1]})
            
            if sample_data:
                st.dataframe(sample_data, use_container_width=True, hide_index=True)
                remaining = len(result1.get("rows", [])) - 5
                if remaining > 0:
                    st.caption(f"... +{remaining} more measurements in Excel")
            
            st.divider()
            
            st.download_button(
                label="⬇️  Download Excel",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")
else:
    st.info("👆 Upload an image to get started")
